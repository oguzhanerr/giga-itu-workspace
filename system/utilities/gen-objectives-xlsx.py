from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/Users/oz/Library/Mobile Documents/iCloud~md~obsidian/Documents/ozv2/projects/self-assessment/Individual Objectives 2026 - Revised.xlsx"

OBJECTIVES = [
    (
        "1",
        "Establish an agile PM framework for the BCN tech team with standardised workflows, tooling, and ceremonies.",
        "Agile process guide, ClickUp dashboards, and ceremony templates.",
        "Framework in active use across all product tracks; team operates it independently.",
    ),
    (
        "2",
        "Build strategic data and technology partnerships to strengthen the Giga product portfolio.",
        "Partner pipeline with active engagement records for MTN, Chargebyte, and one additional partner.",
        "At least two partnerships reach active agreement or signed MoU.",
    ),
    (
        "3",
        "Deliver the product documentation baseline for all active products and scope Phase 2 (Giga Login, Giga Global Data Platform).",
        "PRDs, target audience docs, and UX flows for MST, GigaKIX, and MCVT; v1 PRDs for Phase 2 products.",
        "All products documented to standard; Phase 2 PRDs approved and development-ready.",
    ),
    (
        "4",
        "Embed AI tools into the team's daily ways of working through a structured adoption plan and training.",
        "New Ways of Working document and three training sessions.",
        "AI tooling in active use across at least two team workflows by end of cycle.",
    ),
]

# Gantt: month columns Apr–Dec 2026, start/end month index (0=Apr)
GANTT = [
    ("1 · PM Framework",      0, 8),   # Apr–Dec
    ("2 · Partnerships",       1, 8),   # May–Dec
    ("3 · Product Docs & PRDs",0, 7),   # Apr–Nov
    ("4 · AI Adoption",        2, 8),   # Jun–Dec
]
MONTHS = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# Colours
NAVY   = "1F3864"
BLUE   = "2E75B6"
LIGHT  = "D6E4F0"
GANTT_COLOR = "2E75B6"
GANTT_DONE  = "A9C4E0"
WHITE  = "FFFFFF"
GRAY   = "F2F2F2"

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

wb = Workbook()

# ── Sheet 1: Objectives ──────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Objectives"

ws1.column_dimensions["A"].width = 6
ws1.column_dimensions["B"].width = 52
ws1.column_dimensions["C"].width = 38
ws1.column_dimensions["D"].width = 48

headers = ["#", "Objective", "Deliverable", "Success Criterion"]
for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, color=WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()
ws1.row_dimensions[1].height = 22

for i, (num, obj, deliv, crit) in enumerate(OBJECTIVES, 2):
    fill = PatternFill("solid", fgColor=GRAY if i % 2 == 0 else WHITE)
    for col, val in enumerate([num, obj, deliv, crit], 1):
        cell = ws1.cell(row=i, column=col, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical="top",
                                   horizontal="center" if col == 1 else "left")
        cell.fill = fill
        cell.border = thin_border()
        if col == 1:
            cell.font = Font(bold=True, size=11)
    ws1.row_dimensions[i].height = 72

# ── Sheet 2: Gantt ───────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Gantt")

ws2.column_dimensions["A"].width = 26
for m in range(len(MONTHS)):
    ws2.column_dimensions[get_column_letter(m + 2)].width = 7

# Header row
ws2.cell(row=1, column=1, value="Activity").font = Font(bold=True, color=WHITE, size=11)
ws2.cell(row=1, column=1).fill = PatternFill("solid", fgColor=NAVY)
ws2.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
ws2.cell(row=1, column=1).border = thin_border()
ws2.row_dimensions[1].height = 20

for m, name in enumerate(MONTHS):
    cell = ws2.cell(row=1, column=m + 2, value=name)
    cell.font = Font(bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()

# Activity rows
for r, (label, start, end) in enumerate(GANTT, 2):
    fill_row = PatternFill("solid", fgColor=GRAY if r % 2 == 0 else WHITE)
    name_cell = ws2.cell(row=r, column=1, value=label)
    name_cell.font = Font(bold=True, size=10)
    name_cell.fill = fill_row
    name_cell.alignment = Alignment(vertical="center")
    name_cell.border = thin_border()
    ws2.row_dimensions[r].height = 24

    for m in range(len(MONTHS)):
        cell = ws2.cell(row=r, column=m + 2)
        if start <= m <= end:
            cell.fill = PatternFill("solid", fgColor=GANTT_COLOR)
        else:
            cell.fill = fill_row
        cell.border = thin_border()

wb.save(OUT)
print(f"Saved: {OUT}")
